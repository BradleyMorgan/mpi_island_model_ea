#!/usr/bin/env python

import os
import pandas as pd
import shutil
import re
import openpyxl

head = os.getcwd()

extensions = ('.csv')

pd.set_option('display.max_colwidth', 90)

top_sheet = pd.DataFrame()
sol_sheet = pd.DataFrame()
    
for subdir, dirs, files in os.walk(head):
    #tree = re.match(r'\/?([^\/]+)', relp).groups()
    #tree = re.match(r'\/?([^\/]+)', relp)
    
    print('*** searching subdirectory ' + subdir + ' ***')
    
    for file in files:
    
        print('*** evaluating file ' + file + ' ***')
    
        ext = os.path.splitext(file)[-1].lower()
        
        if ext in extensions and not file.startswith('.') and 'run' not in file and 'solpop' not in file:
            relp = os.path.relpath(subdir, head)
            tree = relp.split("/")
            combo = 'all_' + tree[0] + '.xlsx';
            if os.path.exists(combo):
                writer = pd.ExcelWriter(combo, engine='openpyxl', mode='a')
            else: 
                writer = pd.ExcelWriter(combo, engine='openpyxl')
                s1 = writer.book.create_sheet(title='top' + tree[1])
                s2 = writer.book.create_sheet(title='sol' + tree[1])
                
            print('found matching file: ' + file)
            fn = os.path.join(os.path.relpath(subdir, head), file)
            fout = fn + '.xlsx'
            dcp = head + '/all'
            fcp = dcp + '/' + os.path.split(fout)[1]
            if not os.path.isdir(dcp):
                os.mkdir(dcp, mode = 0o740)
            print('reading: ' + fn)
            df = pd.read_csv(fn)
            #print(df)
            sh = tree[1]
            if 'sol' in file:
                if 'bench' in file:
                    sh = 'bench' + '-sol-' + tree[1]
                if 'evo' in file:
                    sh = 'evo' + '-sol-' + tree[1]
            if 'topo' in file:
                if 'bench' in file:
                    sh = 'bench' + '-topo-' + tree[1]
                if 'evo' in file:
                    sh = 'evo' + '-topo-' + tree[1]
            print('worksheet: ' + sh)
            df.to_excel(writer, sheet_name=sh, header=True, index=False)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            
            if s1.max_column == 1:
                sol1 = df.iloc[:,[0,1]]
                sol1.columns = [ 'run', 'cycle' ]
                #sol1.to_excel(writer, sheet_name='top' + tree[1], header=True, index=False)
                print('created new sheet top' + tree[1])
                sol1 = sol1.reset_index(drop=True)
                top_sheet = pd.concat([top_sheet, sol1], axis=1)
            if s2.max_column == 1:
                sol2 = df.iloc[:,[0,1]]
                sol2.columns = [ 'run', 'cycle' ]
                #sol2.to_excel(writer, sheet_name='sol' + tree[1], header=True, index=False)
                print('created new sheet sol' + tree[1])
                sol2 = sol2.reset_index(drop=True)
                sol_sheet = pd.concat([sol_sheet, sol2], axis=1)
            if 'topo' in file:
                sol1 = df.iloc[:,[3,8]]
                print('sol1 dataframe')
                print(sol1)
                #sol1 = sol1.reset_index(drop=True)
                #top_sheet = top_sheet.reset_index(drop=True)
                if 'bench' in file:
                    print('bench' + file)
                    sol1.columns = [ 'bench_avg_topo_fitness', 'bench_global_best_topo_fitness' ]
                    #sol1.to_excel(writer, sheet_name='top' + tree[1], header=True, index=False)
                    top_sheet = pd.concat([top_sheet, sol1], axis=1)
                if 'evo' in file:
                    print('evo ' + file)
                    sol1.columns = [ 'evo_avg_topo_fitness', 'evo_global_best_topo_fitness' ]
                    #sol1.to_excel(writer, sheet_name='top' + tree[1], header=True, index=False)
                    top_sheet = pd.concat([top_sheet, sol1], axis=1)
            if 'sol' in file:
                sol2 = df.iloc[:,[3,7]]
                print('sol2 dataframe')
                print(sol2)
                #sol2 = sol2.reset_index(drop=True)
                #sol_sheet = sol_sheet.reset_index(drop=True)
                if 'bench' in file:
                    print('bench' + file)
                    sol2.columns = [ 'bench_avg_sol_fitness', 'bench_avg_global_best_sol_fitness' ]
                    #sol2.to_excel(writer, sheet_name='sol' + tree[1], header=True, index=False)
                    sol_sheet = pd.concat([sol_sheet, sol2], axis=1)
                if 'evo' in file:
                    print('evo ' + file)
                    sol2.columns = [ 'evo_avg_sol_fitness', 'evo_global_best_sol_fitness' ]
                    #sol2.to_excel(writer, sheet_name='sol' + tree[1], header=True, index=False)
                    sol_sheet = pd.concat([sol_sheet, sol2], axis=1)
                    
            print('*** top sheet ***')
            print(top_sheet)
            
            print('*** sol sheet ***')
            print(sol_sheet)
            
            top_sheet.to_excel(writer, sheet_name='top' + tree[1], header=True, index=False)
            sol_sheet.to_excel(writer, sheet_name='sol' + tree[1], header=True, index=False)
                
            for name, ws in writer.sheets.items():
                for column in ws.columns:
                    ws.column_dimensions[column[0].column_letter].width = 20
            writer.save()
            


